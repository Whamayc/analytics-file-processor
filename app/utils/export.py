"""Reusable export builder functions shared across pages."""
import io
import re
import math
import datetime
import pandas as pd


def build_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    freeze_header: bool = True,
    autofit: bool = True,
    format_header: bool = True,
    include_header: bool = True,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"
    ws.sheet_view.showGridLines = False

    if format_header:
        HDR_FILL = PatternFill("solid", fgColor="1A1A1A")
        HDR_FONT = Font(color="F59E0B", bold=True, name="Consolas", size=10)
    else:
        HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
        HDR_FONT = Font(bold=True, name="Calibri", size=10)
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center")
    CELL_FONT  = Font(name="Consolas", size=9)
    CELL_ALIGN = Alignment(horizontal="left", vertical="center")
    THIN       = Border(bottom=Side(style="thin", color="DDDDDD"))

    data_start_row = 1
    if include_header:
        for ci, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=str(col_name))
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN; cell.border = THIN
        data_start_row = 2

    for ri, row in enumerate(df.itertuples(index=False), data_start_row):
        for ci, value in enumerate(row, 1):
            if not isinstance(value, str) and pd.isna(value):
                value = None
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = CELL_FONT; cell.alignment = CELL_ALIGN; cell.border = THIN

    if include_header and freeze_header:
        ws.freeze_panes = "A2"

    if autofit:
        for ci, col_name in enumerate(df.columns, 1):
            vals_len = df.iloc[:, ci - 1].astype(str).str.len()
            hdr_len  = len(str(col_name)) if include_header else 0
            max_len  = max(hdr_len, int(vals_len.max()) if len(df) else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _oracle_val(val, dtype) -> str:
    try:
        if val is None or (not isinstance(val, (str, bool)) and pd.isna(val)):
            return "NULL"
    except (TypeError, ValueError):
        pass

    if pd.api.types.is_datetime64_any_dtype(dtype):
        try:
            ts = pd.Timestamp(val)
            if ts.hour == 0 and ts.minute == 0 and ts.second == 0:
                return f"TO_DATE('{ts.strftime('%Y-%m-%d')}', 'YYYY-MM-DD')"
            return f"TO_DATE('{ts.strftime('%Y-%m-%d %H:%M:%S')}', 'YYYY-MM-DD HH24:MI:SS')"
        except Exception:
            pass

    if pd.api.types.is_bool_dtype(dtype):
        return "1" if val else "0"

    if pd.api.types.is_numeric_dtype(dtype):
        try:
            f = float(val)
            if math.isnan(f) or math.isinf(f):
                return "NULL"
            if f == int(f):
                return str(int(f))
            return repr(f)
        except (TypeError, ValueError):
            pass

    s = str(val).replace("'", "''")
    return f"'{s}'"


def _oracle_col(col: str) -> str:
    """Return an Oracle-safe column identifier (uppercase, underscores, max 30 chars)."""
    c = col.strip().upper()
    c = re.sub(r"[^\w]", "_", c)
    c = re.sub(r"_+", "_", c)
    c = c.strip("_")
    return c[:30] or "COL"


def build_oracle_sql(
    df: pd.DataFrame,
    schema: str,
    table: str,
    batch_size: int = 500,
    include_commit: bool = True,
) -> str:
    schema  = re.sub(r"[^\w]", "_", schema.strip().upper()) or "SCHEMA"
    table   = re.sub(r"[^\w]", "_", table.strip().upper())  or "TABLE_NAME"
    cols    = list(df.columns)
    ocols   = [_oracle_col(c) for c in cols]
    col_str = ", ".join(ocols)
    target  = f"{schema}.{table}"

    lines = ["SET DEFINE OFF;", ""]
    n     = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        vals    = [_oracle_val(row[c], df[c].dtype) for c in cols]
        val_str = ", ".join(vals)
        lines.append(f"INSERT INTO {target} ({col_str}) VALUES ({val_str});")
        if batch_size and (i + 1) % batch_size == 0 and i < n - 1:
            lines.append("/")
            if include_commit:
                lines.append("COMMIT;")
            lines.append("")
    lines.append("/")
    if include_commit:
        lines.append("COMMIT;")
    return "\n".join(lines)


def open_outlook_draft(
    recipients: list[str],
    subject: str,
    body: str,
    attachment_bytes: bytes,
    attachment_filename: str,
) -> None:
    """Open a pre-populated Outlook draft for review and manual send.

    Writes the attachment to a temporary file, creates a new Outlook mail
    item via win32com, attaches the file, and calls Display() so the user
    can review and send (or discard) from the Outlook compose window.
    Requires Outlook desktop to be installed and running on the same machine.
    """
    import tempfile
    import os
    import win32com.client  # type: ignore[import]

    # Write attachment bytes to a temp file so Outlook can attach it by path
    suffix = os.path.splitext(attachment_filename)[1] or ".tmp"
    tmp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=suffix,
        prefix="afp_export_",
    )
    try:
        tmp.write(attachment_bytes)
        tmp.flush()
        tmp.close()

        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = olMailItem
        mail.To      = "; ".join(recipients)
        mail.Subject = subject
        mail.Body    = body
        mail.Attachments.Add(tmp.name)
        mail.Display()   # opens compose window — user reviews and clicks Send
    except Exception:
        # Clean up temp file on error; on success Outlook holds it open
        try:
            os.unlink(tmp.name)
        except OSError:
            pass
        raise


def build_pdf(df: pd.DataFrame, source_fname: str, max_rows: int = 2000) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer,
    )
    from reportlab.lib.styles import ParagraphStyle

    AMBER = colors.HexColor("#F59E0B")
    DARK  = colors.HexColor("#0D0D0D")
    DARK2 = colors.HexColor("#141414")
    DARK3 = colors.HexColor("#1A1A1A")
    LIGHT = colors.HexColor("#E5E5E5")
    MUTED = colors.HexColor("#737373")

    n_cols    = len(df.columns)
    page_size = landscape(A4) if n_cols > 6 else A4
    pw        = page_size[0]
    export_dt = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M")
    display   = df.head(max_rows)
    truncated = len(df) > max_rows

    buf = io.BytesIO()

    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#222222"))
        canvas.line(15*mm, 12*mm, pw - 15*mm, 12*mm)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(MUTED)
        canvas.drawString(15*mm, 7*mm, f"Source: {source_fname}")
        canvas.drawCentredString(pw / 2, 7*mm,
                                 "ANALYTICS FILE PROCESSOR — INTERNAL USE ONLY")
        canvas.drawRightString(pw - 15*mm, 7*mm,
                               f"Page {doc.page}  |  {export_dt}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=20*mm,
    )

    SB = ParagraphStyle("SB", fontName="Courier", fontSize=7.5, textColor=MUTED)

    def _tbl_style_base(header_rows=1) -> list:
        return [
            ("BACKGROUND",    (0, 0), (-1, header_rows - 1), DARK3),
            ("TEXTCOLOR",     (0, 0), (-1, header_rows - 1), AMBER),
            ("FONTNAME",      (0, 0), (-1, header_rows - 1), "Courier-Bold"),
            ("FONTSIZE",      (0, 0), (-1, header_rows - 1), 8),
            ("BACKGROUND",    (0, header_rows), (-1, -1),    DARK),
            ("TEXTCOLOR",     (0, header_rows), (-1, -1),    LIGHT),
            ("FONTNAME",      (0, header_rows), (-1, -1),    "Courier"),
            ("FONTSIZE",      (0, header_rows), (-1, -1),    7),
            ("ROWBACKGROUNDS",(0, header_rows), (-1, -1),    [DARK, DARK2]),
            ("GRID",          (0, 0), (-1, -1), 0.25, colors.HexColor("#222")),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ]

    story = []

    if truncated:
        story.append(Paragraph(
            f"Showing first {max_rows:,} of {len(df):,} rows.", SB))
        story.append(Spacer(1, 2*mm))

    col_w = min((pw - 30*mm) / n_cols, 55*mm)
    data_hdr = [str(c) for c in display.columns]
    data_rows = [
        [("" if pd.isna(v) else str(v)) for v in row]
        for _, row in display.iterrows()
    ]
    dt = Table([data_hdr] + data_rows,
               colWidths=[col_w] * n_cols, repeatRows=1)
    dt.setStyle(TableStyle(_tbl_style_base()))
    story.append(dt)

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return buf.getvalue()
